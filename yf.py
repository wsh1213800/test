from openpyxl import load_workbook
wb = load_workbook("12月份衣服销售数据.xlsx")
s=wb.get_sheet_names() #获得表单名字
sheet = wb.get_sheet_by_name(s[0])
r=a=z=x=c=v=b=n=0
for rowNum in range(2,32):
        q=sheet.cell(row=rowNum,column=3).value
        w=sheet.cell(row=rowNum, column=5).value
        if q==253.6:
          z+=w
        elif q==86.3:
          x += w
        elif q==96.8:
          c+=w
        elif q==135.9:
          v+=w
        elif q==65.8:
          b+=w
        else:#49.3
          n+=w
        e=q*w
        r=r+e
        a+=w
print("总销售额为"+str(r))
print("平均每日销售数量为"+str(a/30))
print("羽绒服月销售量占比"+'{:.2%}'.format(z/a))
print("牛仔裤月销售量占比"+'{:.2%}'.format(x/a))
print("风衣月销售量占比"+'{:.2%}'.format(c/a))
print("皮草月销售量占比"+'{:.2%}'.format(v/a))
print("T血月销售量占比"+'{:.2%}'.format(b/a))
print("衬衫月销售量占比"+'{:.2%}'.format(n/a))




